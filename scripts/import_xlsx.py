#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""秋招雷达 · Excel 自动导入脚本。

用法：
  python scripts/import_xlsx.py applications   # 只导入「秋招简历投递.xlsx」
  python scripts/import_xlsx.py zhudi          # 只导入朱迪学姐汇总表
  python scripts/import_xlsx.py                # 两者都导入

依赖：openpyxl（Codex 桌面版自带 Python 已安装）。
数据政策：applications.js 由用户 Excel 生成，勿手改；zhudi.js 为第三方线索层。
"""

import json
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
APPLICATIONS_XLSX = Path(r"C:\Users\Lenovo\Desktop\秋招简历投递.xlsx")
APPLICATIONS_OUT = REPO_ROOT / "data" / "applications.js"
ZHUDI_OUT = REPO_ROOT / "data" / "zhudi.js"


def newest_zhudi_xlsx() -> Path:
    """自动选择最新导出的朱迪学姐汇总表（支持 Downloads 与 E 盘根目录、带 (1) 的重复导出）。"""
    pattern = "27-*朱迪学姐*.xlsx"
    roots = [Path(r"C:\Users\Lenovo\Downloads"), Path(r"E:\\")]
    candidates = []
    for root in roots:
        if root.exists():
            candidates.extend(root.glob(pattern))
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        raise FileNotFoundError(f"未在 Downloads/E 盘根目录找到 {pattern}")
    return candidates[0]


ZHUDI_XLSX = newest_zhudi_xlsx()

COMPANY_MAP = {
    "米哈游": "mihoyo",
    "4399": "game4399",
    "莉莉丝": "lilith",
    "点点": "diandian",
    "字节跳动": "bytedance",
    "大疆": "dji",
    "影石": "insta360",
    "网易雷火": "netease_game",
    "网易互娱": "netease_game",
    "鹰角": "hypergryph",
    "京东": "jd",
    "联想": "lenovo",
    "百度": "baidu",
    "OPPO": "oppo",
    "拼多多": "pinduoduo",
    "阿里千问": "alibaba",
    "阿里灵犀互娱": "alibaba",
    "阿里飞猪": "alibaba",
    "阿里巴巴控股集团": "alibaba",
    "淘宝闪购": "alibaba",
    "阿里云": "alibaba",
    "阿里国际数字商业集团": "alibaba",
    "Token Foundry": "alibaba",
    "高德地图": "alibaba",
    "菜鸟": "alibaba",
    "虎鲸文娱集团": "alibaba",
    "盒马": "alibaba",
    "阿里健康": "alibaba",
    "b站": "bilibili",
    "G社": "garena",
    "滴滴": "didi",
    "快手": "kuaishou",
    "小鹏": "xiaopeng",
    "顺丰": "shunfeng",
    "得物": "dewu",
    "盛趣游戏": "shengqu",
    "去哪儿旅行": "qunar",
    "巨人网络": "ztgame",
    "小米": "xiaomi",
    "蚂蚁集团": "antgroup",
    "腾讯": "tencent",
    "理想": "lixiang",
    "的物": "dewu",
    "掌趣": "zhangqu",
    "作业帮": "zuoyebang",
    "迅雷": "xunlei",
    "沐瞳": "moonton",
    "合合": "intsig",
    "搜狐": "cyou",
    "地平线": "horizon",
    "拓竹": "bambulab",
    "安可创新": "anker",
    "shopee": "shopee",
    "远景": "envision",
    "吉利": "geely",
    "韶音科技": "shokz",
    "招商云创": "mbcloud",
    "科大讯飞": "iflytek",
    "亿联": "yealink",
    "美团": "meituan",
    "传音控股": "transsion",
    "TP-Link": "tplink",
    "海信": "hisense",
    "卓驭": "zhuoyu",
    "普渡": "pudu",
    "wind": "wind",
    "宁德时代": "catl",
    "蔚来": "nio",
    "九号公司": "ninebot",
    "凡岛": "fandow",
    "原力": "dexmal",
    "嘉伯士": "carlsberg",
    "携程": "ctrip",
    "库洛": "kuro",
    "途游": "tuyoo",
    "深蓝互动": "bluepoch",
    "千里科技": "qianli",
    "经纬恒润": "hirain",
    "巨一科技": "jee",
    "商汤": "sensetime",
    "禾赛经济": "hesai",
    "四方": "sifang",
    "李森": "lisen",
    "黑白调": "heibaidiao",
    "养生堂农夫山泉": "nongfu",
    "美的": "midea",
    "亿道": "emdoor",
    "正浩创新": "ecoflow",
    "meta app": "metaapp",
    "雷赛": "leadshine",
    "思摩尔国际": "smoore",
    "雷霆吉比特": "gbits",
}


def pad(n: int) -> str:
    return str(n).zfill(2)


def excel_to_iso(v):
    if isinstance(v, datetime):
        return f"{v.year}-{pad(v.month)}-{pad(v.day)}"
    if isinstance(v, (int, float)):
        d = datetime(1899, 12, 31) + timedelta(days=float(v))
        return f"{d.year}-{pad(d.month)}-{pad(d.day)}"
    return ""


def norm_date(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return excel_to_iso(v)
    if isinstance(v, (int, float)):
        return excel_to_iso(v)
    s = str(v).strip()
    m = re.match(r"^(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})", s)
    if m:
        return f"{m.group(1)}-{pad(int(m.group(2)))}-{pad(int(m.group(3)))}"
    return s[:40]


def text(v):
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, (list, tuple)):
        return "、".join(text(x) for x in v)
    return str(v)


def trunc(s: str, n: int) -> str:
    return s if len(s) <= n else s[:n] + "…"


def cell(row, idx):
    if idx is None or row is None or idx >= len(row):
        return None
    return row[idx]


def header_cols(headers):
    """按表头名称找列，兼容朱迪导出不同版本的列顺序。"""
    h = [text(x) for x in headers]

    def find(*needles):
        for i, name in enumerate(h):
            if not name:
                continue
            for n in needles:
                if n in name:
                    return i
        return None

    return {
        "company": find("公司"),
        "industry": find("行业类别", "行业"),
        "batch": find("批次"),
        "grad": find("届次"),
        "locations": find("工作地点", "地点"),
        "positions": find("招聘岗位", "岗位"),
        "startDate": find("开始时间", "开始"),
        "endDate": find("截止时间", "结束时间", "截止", "结束"),
        "applyUrl": find("简历投递链接", "投递链接"),
        "announceUrl": find("公告链接", "原文链接"),
        "exam": find("是否笔试", "笔试"),
    }


def import_applications() -> int:
    wb = openpyxl.load_workbook(APPLICATIONS_XLSX, data_only=True)
    sheet = wb.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    apps = []
    for i in range(1, len(rows)):
        row = rows[i]
        company = text(cell(row, 0))
        position = text(cell(row, 1))
        if not company or not position:
            continue
        stage, note = normalize_stage(cell(row, 3))
        apps.append(
            {
                "companyId": COMPANY_MAP.get(company, ""),
                "companyName": company,
                "position": position,
                "appliedAt": excel_to_iso(cell(row, 2)),
                "stage": stage,
                "note": note,
            }
        )
    header = "// 自动从「秋招简历投递.xlsx」生成，请勿手改；用户更新表格后重新运行导入脚本\n"
    body = "window.QIUZHAO_APPLICATIONS = " + json.dumps(apps, ensure_ascii=False, indent=2) + ";\n"
    APPLICATIONS_OUT.write_text(header + body, encoding="utf-8")
    print(f"applications: imported {len(apps)} rows -> {APPLICATIONS_OUT.name}")
    return len(apps)


_CN_DIGIT = {"零": 0, "一": 1, "二": 2, "两": 2, "三": 3, "四": 4, "五": 5, "六": 6, "七": 7, "八": 8, "九": 9}
_KNOWN_STAGES = {"已投递", "笔试", "面试", "Offer", "已挂"}


def _cn_to_int(s: str):
    """中文数字 1-31 转整数：八->8 十->10 十二->12 二十->20 二十一->21"""
    if not s:
        return None
    if s.isdigit():
        return int(s)
    total = 0
    for i, ch in enumerate(s):
        if ch == "十":
            if total == 0:
                total = 10
            elif total < 10:
                total *= 10
        elif ch in _CN_DIGIT:
            v = _CN_DIGIT[ch]
            nxt = s[i + 1] if i + 1 < len(s) else ""
            if nxt == "十":
                total = v * 10
            elif total == 10 and i > 0 and s[i - 1] == "十":
                total += v
            else:
                total += v
        else:
            return None
    return total


def _normalize_date_note(s: str) -> str:
    """「八月22号挂」->「8/22挂」；「9月一号挂」->「9/1挂」；无法解析则原样返回"""
    import re

    s = s.replace("号", "").replace("日", "")
    m = re.search(r"(\d+|[一二三四五六七八九十两]+)月(\d+|[一二三四五六七八九十两]+)", s)
    if m:
        mo = _cn_to_int(m.group(1))
        day = _cn_to_int(m.group(2))
        if mo is not None and day is not None:
            return f"{mo}/{day}挂"
    return s


def normalize_stage(raw) -> tuple:
    """Excel「面试」列 -> (stage, note)。自由文本含「挂」归为已挂并保留日期备注。"""
    s = text(raw).strip()
    if not s:
        return "已投递", ""
    clean = s.strip(" '\"'“‘”’")
    if clean in _KNOWN_STAGES:
        return clean, ""
    if "挂" in s:
        return "已挂", _normalize_date_note(s)
    if "面试" in s:
        return "面试", s
    return s, ""


def import_zhudi() -> tuple:
    wb = openpyxl.load_workbook(ZHUDI_XLSX, data_only=True)
    main = next(
        (ws for ws in wb.worksheets if "校招汇总表" in ws.title),
        wb.worksheets[0],
    )
    codes_sheet = next(
        (ws for ws in wb.worksheets if "内推码" in ws.title),
        wb.worksheets[4] if len(wb.worksheets) > 4 else None,
    )
    rows = list(main.iter_rows(values_only=True))
    code_rows = list(codes_sheet.iter_rows(values_only=True)) if codes_sheet else []
    cols = header_cols(rows[0]) if rows else {}

    out_rows = []
    for i in range(1, len(rows)):
        r = rows[i]
        company = text(cell(r, cols["company"]))
        if not company or "（必看）" in company:
            continue
        grad = text(cell(r, cols["grad"]))
        if "2027" not in grad:
            continue
        out_rows.append(
            {
                "company": trunc(company, 60),
                "industry": trunc(text(cell(r, cols["industry"])), 40),
                "batch": trunc(text(cell(r, cols["batch"])), 20),
                "grad": trunc(grad, 40),
                "locations": trunc(text(cell(r, cols["locations"])), 90),
                "positions": trunc(text(cell(r, cols["positions"])), 160),
                "startDate": norm_date(cell(r, cols["startDate"])),
                "endDate": norm_date(cell(r, cols["endDate"])),
                "applyUrl": text(cell(r, cols["applyUrl"]))[:300],
                "announceUrl": text(cell(r, cols["announceUrl"]))[:300],
                "exam": trunc(text(cell(r, cols["exam"])), 20),
            }
        )

    codes = []
    for i in range(1, len(code_rows)):
        r = code_rows[i]
        company = text(cell(r, 0))
        code = text(cell(r, 1))
        if company and code:
            codes.append({"company": trunc(company, 60), "code": code[:60]})

    data = {
        "updatedAt": datetime.now().strftime("%Y-%m-%d"),
        "source": "朱迪学姐汇总表（第三方整理）",
        "sourceUrl": "https://acnr1ayjzqxf.feishu.cn/base/ISNobVuXAagJBFssvszcYeqQnfb",
        "rows": out_rows,
        "codes": codes,
    }
    header = (
        "// 自动生成：朱迪学姐秋招汇总表（2027届 + 内推码）\n"
        "// 来源为第三方整理，属线索层；投递前请以官方公告为准\n"
    )
    body = "window.QIUZHAO_ZHUDI = " + json.dumps(data, ensure_ascii=False, indent=1) + ";\n"
    ZHUDI_OUT.write_text(header + body, encoding="utf-8")
    if codes_sheet is None:
        print("zhudi: WARNING 未找到「内推码」sheet，内推码保持为空")
    print(f"zhudi: rows={len(out_rows)} codes={len(codes)} -> {ZHUDI_OUT.name}")
    return len(out_rows), len(codes)


def main():
    targets = sys.argv[1:] or ["applications", "zhudi"]
    for t in targets:
        if t == "applications":
            import_applications()
        elif t == "zhudi":
            import_zhudi()
        else:
            print(f"unknown target: {t}", file=sys.stderr)
            sys.exit(2)


if __name__ == "__main__":
    main()
